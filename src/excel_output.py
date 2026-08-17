from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from datetime import date
from typing import cast
import math
import os

COLUMN_WIDTH_PADDING = 2

FIELD_ORDER = [
    "group_name",
    "Configuracion",
    "AvailableOnHandQuantity",
    "avg_daily_consumption_7d",
    "days_remaining_7d",
    "avg_daily_consumption_30d",
    "days_remaining_30d",
]

HEADER_LABELS = {
    "group_name": "Código de Artículo",
    "Configuracion": "Configuración",
    "AvailableOnHandQuantity": "Inventario Disponible (kg)",
    "avg_daily_consumption_7d": "Consumo Diario Promedio (kg/día) - Últimos 7 días",
    "days_remaining_7d": "Cobertura Restante (días) - Últimos 7 días",
    "avg_daily_consumption_30d": "Consumo Diario Promedio (kg/día) - Últimos 30 días",
    "days_remaining_30d": "Cobertura Restante (días) - Últimos 30 días",
}

DAYS_REMAINING_FIELDS = ("days_remaining_7d", "days_remaining_30d")
SUM_FIELDS = ["AvailableOnHandQuantity", "avg_daily_consumption_7d", "avg_daily_consumption_30d"]
LOW_COVERAGE_THRESHOLD_DAYS = 60


def setup_sheet(headers: list):
    wb = Workbook()
    ws = cast(Worksheet, wb.active)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="21295C", end_color="21295C", fill_type="solid")

    today = date.today()
    title = f'Reporte de Cobertura de Materia Prima - {today.strftime("%d/%m/%Y")}'
    ws['A1'] = title

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    for col in range(len(headers)):
        cell = ws.cell(row=3, column=col + 1)
        cell.value = headers[col]
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[get_column_letter(col + 1)].width = len(str(headers[col])) + COLUMN_WIDTH_PADDING

    return wb, ws


def write_data_row(ws, row_number, group_data):
    LOW_COVERAGE_FILL = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

    for col_num, key in enumerate(FIELD_ORDER):
        value = group_data[key]
        cell = ws.cell(row=row_number, column=col_num + 1)
        if value is None or (isinstance(value, float) and math.isnan(value)):
            cell.value = "N/A"
        else:
            cell.value = value
            if isinstance(value, (int, float)):
                cell.number_format = '#,##0' if key in DAYS_REMAINING_FIELDS else '#,##0.00'
                if key in DAYS_REMAINING_FIELDS and value <= LOW_COVERAGE_THRESHOLD_DAYS:
                    cell.fill = LOW_COVERAGE_FILL


def write_totals_row(ws, row_number, all_groups_data):
    for col_num, key in enumerate(FIELD_ORDER):
        cell = ws.cell(row=row_number, column=col_num + 1)
        if key in SUM_FIELDS:
            values = [row[key] for row in all_groups_data]
            total = sum(values)
            cell.value = total
            cell.number_format = '#,##0.00'
            cell.font = Font(bold=True)
        elif key == "group_name":
            cell.value = "Total:"
            cell.font = Font(bold=True)


def write_unclassified_note(ws, row_number, unclassified_items, total_columns):
    if not unclassified_items:
        return

    entries = [
        f"{item} ({config})" if config else f"{item}"
        for item, config in unclassified_items
    ]
    note = "Artículos no clasificados (no incluidos en este reporte): " + ", ".join(entries)

    cell = ws.cell(row=row_number, column=1)
    cell.value = note
    cell.font = Font(italic=True)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=total_columns)
    ws.row_dimensions[row_number].height = 60


def generate_report(report_data, unclassified_items=None):
    filename = f"cobertura_mp_{date.today().isoformat()}.xlsx"
    filepath = os.path.join(os.getcwd(), filename)

    headers = [HEADER_LABELS[field] for field in FIELD_ORDER]
    wb, ws = setup_sheet(headers)

    records = report_data.to_dict('records')

    row_num = 4
    for record in records:
        write_data_row(ws, row_num, record)
        row_num += 1

    write_totals_row(ws, row_num, records)
    write_unclassified_note(ws, row_num + 2, unclassified_items or [], len(FIELD_ORDER))

    wb.save(filepath)
    return filepath
